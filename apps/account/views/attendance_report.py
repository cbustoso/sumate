import re
import pandas as pd
from django.http import HttpResponse
from django.utils.timezone import make_naive, now
from django.db.models import Sum, OuterRef, Subquery
from openpyxl.styles import NamedStyle, Font, PatternFill
from openpyxl.utils import get_column_letter
from apps.attendance.models import Attendance
from apps.account.models import Career
from apps.semester.models import Semester
import io

def generate_attendee_excel(request, career_id, semester_id):
    filename = generate_filename(career_id, semester_id)
    attendances = get_attendances(career_id, semester_id)
    df = prepare_dataframe(attendances)
    output = create_excel_file_in_memory(df)
    return send_excel_file_as_http_response(output, filename)

def generate_filename(career_id, semester_id):
    career_name = Career.objects.get(pk=career_id).name
    semester_name = Semester.objects.get(pk=semester_id).name
    clean_career_name = clean_filename(career_name)
    clean_semester_name = clean_filename(semester_name)
    date_str = now().date().isoformat()
    return f"{date_str}_{clean_career_name}_{clean_semester_name}.xlsx"

def clean_filename(name):
    return re.sub(r'[^\w\s-]', '', name).strip().replace(' ', '_')

def get_attendances(career_id, semester_id):
    points_sum_subquery = Attendance.objects.filter(
        attendee=OuterRef('attendee'),
        event__semester_id=semester_id
    ).values('attendee').annotate(total=Sum('points')).values('total')[:1]

    query = Attendance.objects.filter(
        attendee__career_id=career_id,
        event__semester_id=semester_id
    ).values(
        'attendee__run', 'attendee__first_name', 'attendee__last_name', 'attendee__email',
        'event__name', 'event__event_type__name', 'event__semester__name', 'created_at', 'points'
    ).annotate(
        total_points=Subquery(points_sum_subquery)
    ).order_by('attendee__run')

    return list(query) if query.exists() else []

def prepare_dataframe(attendances):
    if not attendances:
        columns = [
            'RUN', 'Nombre', 'Apellido', 'Correo electrónico', 'Nombre del evento',
            'Tipo de evento', 'Nombre del semestre', 'Fecha de asistencia',
            'Puntos otorgados', 'Puntos totales', 'Nombre de la carrera'
        ]
        return pd.DataFrame(columns=columns)

    df = pd.DataFrame(attendances)
    df['created_at'] = df['created_at'].apply(lambda x: make_naive(x))
    
    df.rename(columns={
        'attendee__run': 'RUN',
        'attendee__first_name': 'Nombre',
        'attendee__last_name': 'Apellido',
        'attendee__email': 'Correo electrónico',
        'event__name': 'Nombre del evento',
        'event__event_type__name': 'Tipo de evento',
        'event__semester__name': 'Nombre del semestre',
        'created_at': 'Fecha de asistencia',
        'points': 'Puntos otorgados',
        'total_points': 'Puntos totales',
        'attendee__career__name': 'Nombre de la carrera'
    }, inplace=True)
    return df

def create_excel_file_in_memory(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Asistencias')
        apply_excel_styles(writer)
    output.seek(0)
    return output

def apply_excel_styles(writer):
    workbook = writer.book
    worksheet = writer.sheets['Asistencias']
    header_style = NamedStyle(name="header_style")
    header_style.font = Font(bold=True, color="FFFFFF")
    header_style.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    workbook.add_named_style(header_style)
    for cell in worksheet["1:1"]:
        cell.style = header_style
    adjust_column_widths(worksheet)

def adjust_column_widths(worksheet):
    for column_cells in worksheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length

def send_excel_file_as_http_response(output, filename):
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
